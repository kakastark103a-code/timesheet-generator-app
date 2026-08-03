import os
import io
import re
import zipfile
import calendar
from datetime import datetime, date
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from timesheet_generator import (
    get_working_days,
    generate_domain_timesheet,
    find_template_path,
    get_singapore_holidays,
    scan_available_domain_templates,
    extract_resources_from_summary,
    extract_resources_from_timesheet,
    parse_comment_notes,
    DOMAIN_FILE_MAP,
    DOMAIN_NAMES
)
from ai_assistant import audit_timesheet_data, process_ai_chat_command, apply_fixes_to_workbook
import openpyxl

import tempfile

app = Flask(__name__, template_folder='templates', static_folder='static')
TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'timesheets_extracted')
OUTPUT_DIR = os.path.join(tempfile.gettempdir(), 'generated_timesheets')

# Ensure writable directories exist on cold start (Vercel /tmp is writable)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# In-memory store for custom members per domain
CUSTOM_MEMBERS_STORE = {}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/domains', methods=['GET'])
def get_domains():
    available_domains = scan_available_domain_templates(TEMPLATE_DIR)
    domains_list = []
    
    predefined_keys = list(DOMAIN_FILE_MAP.keys())
    for key in predefined_keys:
        if key in available_domains:
            info = available_domains[key]
            domains_list.append({
                'key': key,
                'name': info['name'],
                'template': info['filename'],
                'available': True,
                'is_custom': False
            })
            
    for key, info in available_domains.items():
        if key not in predefined_keys:
            domains_list.append({
                'key': key,
                'name': info['name'],
                'template': info['filename'],
                'available': True,
                'is_custom': True
            })
            
    return jsonify({'domains': domains_list})

@app.route('/api/members', methods=['GET', 'POST'])
def handle_members():
    if request.method == 'GET':
        domain_key = request.args.get('domain', 'identity')
        if domain_key in CUSTOM_MEMBERS_STORE:
            return jsonify({'domain': domain_key, 'members': CUSTOM_MEMBERS_STORE[domain_key]})
            
        template_path = find_template_path(domain_key, TEMPLATE_DIR)
        if not template_path:
            return jsonify({'domain': domain_key, 'members': []})
            
        try:
            wb = openpyxl.load_workbook(template_path, data_only=True)
            members = extract_resources_from_summary(wb)
            if not members:
                members = extract_resources_from_timesheet(wb['Timesheet']) if 'Timesheet' in wb.sheetnames else []
            wb.close()
            return jsonify({'domain': domain_key, 'members': members})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
            
    elif request.method == 'POST':
        data = request.json or {}
        domain_key = data.get('domain')
        members = data.get('members', [])
        if not domain_key:
            return jsonify({'error': 'Missing domain key'}), 400
            
        CUSTOM_MEMBERS_STORE[domain_key] = members
        return jsonify({'message': f'Updated members for domain "{domain_key}"', 'count': len(members)})

@app.route('/api/upload', methods=['POST'])
def upload_template():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
        
    uploaded_file = request.files['file']
    if not uploaded_file or uploaded_file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
        
    filename = secure_filename(uploaded_file.filename)
    if not filename.endswith('.xlsx'):
        return jsonify({'error': 'Only Excel files (.xlsx) are allowed as template'}), 400
        
    save_path = os.path.join(TEMPLATE_DIR, filename)
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    uploaded_file.save(save_path)
    
    available_domains = scan_available_domain_templates(TEMPLATE_DIR)
    return jsonify({
        'message': f'Template "{filename}" uploaded successfully!',
        'filename': filename,
        'domains_count': len(available_domains)
    })

@app.route('/api/singapore-holidays', methods=['GET'])
def get_sg_holidays():
    year_str = request.args.get('year', str(datetime.now().year))
    month_str = request.args.get('month')
    try:
        year = int(year_str)
        month = int(month_str) if month_str else None
    except ValueError:
        return jsonify({'error': 'Invalid year or month format'}), 400

    hols = get_singapore_holidays(year, month)
    hols_list = [{'date': d, 'name': name} for d, name in sorted(hols.items())]
    return jsonify({'year': year, 'month': month, 'holidays': hols_list})

@app.route('/api/preview', methods=['POST'])
def preview_month():
    data = request.json or {}
    month_str = data.get('month', datetime.now().strftime('%Y-%m'))
    public_holidays = data.get('public_holidays', [])
    include_sg_holidays = data.get('auto_sg_holidays', True)
    
    try:
        dt = datetime.strptime(month_str, '%Y-%m')
        year, month = dt.year, dt.month
    except ValueError:
        return jsonify({'error': 'Invalid month format, expected YYYY-MM'}), 400

    _, total_days = calendar.monthrange(year, month)
    working_days = get_working_days(year, month)
    
    ph_dates = set()
    ph_names = {}

    if include_sg_holidays:
        sg_hols = get_singapore_holidays(year, month)
        for d_str, name in sg_hols.items():
            dt_obj = datetime.strptime(d_str, '%Y-%m-%d').date()
            ph_dates.add(dt_obj)
            ph_names[dt_obj] = name

    for ph in public_holidays:
        try:
            dt_obj = datetime.strptime(ph.strip(), '%Y-%m-%d').date()
            ph_dates.add(dt_obj)
            if dt_obj not in ph_names:
                ph_names[dt_obj] = "Public Holiday"
        except ValueError:
            pass

    day_list = []
    for day in range(1, total_days + 1):
        cur_date = date(year, month, day)
        is_weekend = cur_date.weekday() >= 5
        is_ph = cur_date in ph_dates
        day_list.append({
            'date': cur_date.strftime('%Y-%m-%d'),
            'day_num': day,
            'weekday_name': cur_date.strftime('%a'),
            'is_weekend': is_weekend,
            'is_working': not is_weekend,
            'is_holiday': is_ph,
            'holiday_name': ph_names.get(cur_date, "") if is_ph else ""
        })
        
    working_days_count = len(working_days)
    month_name = datetime(year, month, 1).strftime('%B %Y')
    
    return jsonify({
        'year': year,
        'month': month,
        'month_name': month_name,
        'total_days': total_days,
        'working_days_count': working_days_count,
        'weekend_days_count': total_days - working_days_count,
        'holidays_count': len(ph_dates),
        'days': day_list
    })

@app.route('/api/review', methods=['POST'])
def review_timesheet():
    data = request.json or {}
    month_str = data.get('month', datetime.now().strftime('%Y-%m'))
    domain_key = data.get('domain', 'identity')
    comment_notes = data.get('comment_notes', '')
    public_holidays = data.get('public_holidays', [])
    include_sg_holidays = data.get('auto_sg_holidays', True)
    custom_members_input = data.get('custom_members', {})

    try:
        dt = datetime.strptime(month_str, '%Y-%m')
        year, month = dt.year, dt.month
    except ValueError:
        return jsonify({'error': 'Invalid month format'}), 400

    template_path = find_template_path(domain_key, TEMPLATE_DIR)
    if not template_path:
        return jsonify({'error': f'Template not found for domain {domain_key}'}), 404

    temp_out = os.path.join(OUTPUT_DIR, f"_temp_review_{domain_key}.xlsx")
    domain_members = custom_members_input.get(domain_key) or CUSTOM_MEMBERS_STORE.get(domain_key)

    generate_domain_timesheet(
        template_path=template_path,
        year=year,
        month=month,
        output_path=temp_out,
        public_holidays=public_holidays,
        include_sg_holidays=include_sg_holidays,
        custom_members=domain_members,
        comment_notes=comment_notes,
        domain_key=domain_key
    )

    wb = openpyxl.load_workbook(temp_out, data_only=True)
    
    summary_data = []
    if 'Summary' in wb.sheetnames:
        sheet = wb['Summary']
        for r in range(2, sheet.max_row + 1):
            name = sheet.cell(r, 2).value
            if name:
                summary_data.append({
                    'no': sheet.cell(r, 1).value,
                    'name': str(name),
                    'team': sheet.cell(r, 3).value,
                    'location': sheet.cell(r, 5).value,
                    'working_days': sheet.cell(r, 6).value,
                    'weekend_ot': sheet.cell(r, 7).value,
                    'weekday_ot': sheet.cell(r, 8).value,
                    'ph_ot': sheet.cell(r, 9).value,
                    'total_ot': sheet.cell(r, 10).value,
                    'leaves': sheet.cell(r, 11).value
                })

    balance_data = []
    bal_sheet_name = 'Balance Leave' if 'Balance Leave' in wb.sheetnames else ('Leave Balance' if 'Leave Balance' in wb.sheetnames else None)
    if bal_sheet_name:
        sheet = wb[bal_sheet_name]
        for r in range(3, sheet.max_row + 1):
            name = sheet.cell(r, 2).value
            if name:
                balance_data.append({
                    'name': str(name),
                    'total_leave': sheet.cell(r, 6).value,
                    'balance_upto': sheet.cell(r, 7).value,
                    'balance_in_month': sheet.cell(r, 10).value
                })

    timesheet_data = []
    if 'Timesheet' in wb.sheetnames:
        sheet = wb['Timesheet']
        for r in range(2, min(sheet.max_row + 1, 100)):
            name = sheet.cell(r, 7).value
            if name:
                dt_val = sheet.cell(r, 2).value
                if isinstance(dt_val, (datetime, date)):
                    dt_str = dt_val.strftime('%Y-%m-%d')
                else:
                    dt_str = str(dt_val)
                timesheet_data.append({
                    'date': dt_str,
                    'work_item_type': sheet.cell(r, 3).value,
                    'name': str(name),
                    'task': sheet.cell(r, 6).value,
                    'hours': sheet.cell(r, 10).value
                })

    wb.close()
    if os.path.exists(temp_out):
        try: os.remove(temp_out)
        except Exception: pass

    anomalies = audit_timesheet_data(summary_data, balance_data, timesheet_data, month_str)

    return jsonify({
        'domain': domain_key,
        'month': month_str,
        'summary': summary_data,
        'balance': balance_data,
        'timesheet_sample': timesheet_data[:30],
        'anomalies': anomalies
    })

@app.route('/api/ai-chat', methods=['POST'])
def ai_chat():
    data = request.json or {}
    prompt = data.get('prompt', '')
    current_notes = data.get('comment_notes', '')
    domain_key = data.get('domain', 'identity')
    api_key = data.get('api_key')
    base_url = data.get('base_url')

    if not prompt:
        return jsonify({'error': 'Missing prompt'}), 400

    result = process_ai_chat_command(prompt, current_notes, domain_key, api_key, base_url)
    return jsonify(result)

@app.route('/api/review-multi-upload', methods=['POST'])
def review_multi_uploaded_files():
    uploaded_files = request.files.getlist('files')
    if not uploaded_files:
        if 'file' in request.files:
            uploaded_files = [request.files['file']]
            
    if not uploaded_files or len(uploaded_files) == 0:
        return jsonify({'error': 'No files uploaded'}), 400

    reports = []
    batch_prefix = f"batch_{int(datetime.now().timestamp())}"

    for ufile in uploaded_files:
        if not ufile or ufile.filename == '':
            continue
            
        filename = secure_filename(ufile.filename)
        if not filename.endswith('.xlsx'):
            continue

        file_token = f"{batch_prefix}_{filename}"
        temp_path = os.path.join(OUTPUT_DIR, f"_uploaded_{file_token}")
        ufile.save(temp_path)

        try:
            wb = openpyxl.load_workbook(temp_path, data_only=True)
            summary_data = []
            if 'Summary' in wb.sheetnames:
                sheet = wb['Summary']
                for r in range(2, sheet.max_row + 1):
                    name = sheet.cell(r, 2).value
                    if name:
                        summary_data.append({
                            'no': sheet.cell(r, 1).value,
                            'name': str(name),
                            'team': sheet.cell(r, 3).value,
                            'location': sheet.cell(r, 5).value,
                            'working_days': sheet.cell(r, 6).value,
                            'weekend_ot': sheet.cell(r, 7).value,
                            'weekday_ot': sheet.cell(r, 8).value,
                            'ph_ot': sheet.cell(r, 9).value,
                            'total_ot': sheet.cell(r, 10).value,
                            'leaves': sheet.cell(r, 11).value
                        })

            balance_data = []
            bal_sheet_name = 'Balance Leave' if 'Balance Leave' in wb.sheetnames else ('Leave Balance' if 'Leave Balance' in wb.sheetnames else None)
            if bal_sheet_name:
                sheet = wb[bal_sheet_name]
                for r in range(3, sheet.max_row + 1):
                    name = sheet.cell(r, 2).value
                    if name:
                        balance_data.append({
                            'name': str(name),
                            'total_leave': sheet.cell(r, 6).value,
                            'balance_upto': sheet.cell(r, 7).value,
                            'balance_in_month': sheet.cell(r, 10).value
                        })

            timesheet_data = []
            if 'Timesheet' in wb.sheetnames:
                sheet = wb['Timesheet']
                for r in range(2, sheet.max_row + 1):
                    name = sheet.cell(r, 7).value
                    if name:
                        dt_val = sheet.cell(r, 2).value
                        dt_str = dt_val.strftime('%Y-%m-%d') if isinstance(dt_val, (datetime, date)) else str(dt_val)
                        timesheet_data.append({
                            'date': dt_str,
                            'work_item_type': sheet.cell(r, 3).value,
                            'name': str(name),
                            'task': sheet.cell(r, 6).value,
                            'hours': sheet.cell(r, 10).value
                        })

            wb.close()
            anomalies = audit_timesheet_data(summary_data, balance_data, timesheet_data, "Current")

            reports.append({
                'file_token': file_token,
                'filename': filename,
                'summary': summary_data,
                'balance': balance_data,
                'timesheet_sample': timesheet_data[:50],
                'anomalies': anomalies
            })
        except Exception as e:
            reports.append({
                'file_token': file_token,
                'filename': filename,
                'error': f"Lỗi đọc file: {str(e)}",
                'anomalies': []
            })

    return jsonify({'reports': reports})

@app.route('/api/apply-fixes-download', methods=['POST'])
def apply_fixes_and_download():
    payload = request.json or {}
    approved_files = payload.get('approved_files', [])
    
    if not approved_files:
        return jsonify({'error': 'No approved fixes provided'}), 400

    fixed_files_paths = []

    for item in approved_files:
        file_token = item.get('file_token')
        fix_actions = item.get('fix_actions', [])
        filename = item.get('filename', 'Timesheet_Fixed.xlsx')

        src_path = os.path.join(OUTPUT_DIR, f"_uploaded_{file_token}")
        if not os.path.exists(src_path):
            continue

        try:
            wb = openpyxl.load_workbook(src_path)
            apply_fixes_to_workbook(wb, fix_actions)
            
            out_filename = f"Fixed_{filename}"
            out_path = os.path.join(OUTPUT_DIR, f"_result_{out_filename}")
            wb.save(out_path)
            wb.close()

            fixed_files_paths.append((out_filename, out_path))
        except Exception as e:
            print(f"Error applying fixes to {file_token}: {str(e)}")

    if not fixed_files_paths:
        return jsonify({'error': 'Failed to process any fixed files'}), 500

    # Clean up original uploads asynchronously
    for item in approved_files:
        file_token = item.get('file_token')
        src_path = os.path.join(OUTPUT_DIR, f"_uploaded_{file_token}")
        if os.path.exists(src_path):
            try: os.remove(src_path)
            except Exception: pass

    # Single File Return
    if len(fixed_files_paths) == 1:
        out_filename, out_path = fixed_files_paths[0]
        return send_file(
            out_path,
            as_attachment=True,
            download_name=out_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # Multiple Files Return as ZIP
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for out_filename, out_path in fixed_files_paths:
            zf.write(out_path, arcname=out_filename)
            try: os.remove(out_path)
            except Exception: pass

    memory_file.seek(0)
    return send_file(
        memory_file,
        as_attachment=True,
        download_name=f"Fixed_Timesheets_Batch_{datetime.now().strftime('%Y%m%d')}.zip",
        mimetype='application/zip'
    )

@app.route('/api/generate', methods=['POST'])
def generate_timesheets():
    data = request.json or {}
    month_str = data.get('month', datetime.now().strftime('%Y-%m'))
    selected_domains = data.get('domains', ['all'])
    public_holidays = data.get('public_holidays', [])
    include_sg_holidays = data.get('auto_sg_holidays', True)
    comment_notes = data.get('comment_notes', "")
    custom_members_input = data.get('custom_members', {})
    
    try:
        dt = datetime.strptime(month_str, '%Y-%m')
        year, month = dt.year, dt.month
    except ValueError:
        return jsonify({'error': 'Invalid month format, expected YYYY-MM'}), 400
        
    month_suffix = datetime(year, month, 1).strftime('%b%Y')
    available_domains = scan_available_domain_templates(TEMPLATE_DIR)
    
    if 'all' in selected_domains or not selected_domains:
        target_keys = list(available_domains.keys())
    else:
        target_keys = selected_domains

    generated_files = []
    
    for key in target_keys:
        template_path = find_template_path(key, TEMPLATE_DIR)
        if not template_path:
            continue
            
        domain_info = available_domains.get(key, {})
        domain_name = domain_info.get('name', key.upper())
        out_filename = f"FPT_{domain_name}_Timesheet_{month_suffix}.xlsx"
        out_filepath = os.path.join(OUTPUT_DIR, out_filename)
        
        domain_members = custom_members_input.get(key) or CUSTOM_MEMBERS_STORE.get(key)
        
        generate_domain_timesheet(
            template_path=template_path,
            year=year,
            month=month,
            output_path=out_filepath,
            public_holidays=public_holidays,
            include_sg_holidays=include_sg_holidays,
            custom_members=domain_members,
            comment_notes=comment_notes,
            domain_key=key
        )
        generated_files.append((out_filename, out_filepath))

    if not generated_files:
        return jsonify({'error': 'No valid template found to generate timesheet.'}), 404

    if len(generated_files) == 1:
        fname, fpath = generated_files[0]
        return send_file(
            fpath,
            as_attachment=True,
            download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    zip_filename = f"FPT_Timesheets_{month_suffix}.zip"
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname, fpath in generated_files:
            zf.write(fpath, arcname=fname)
    memory_file.seek(0)

    return send_file(
        memory_file,
        as_attachment=True,
        download_name=zip_filename,
        mimetype='application/zip'
    )

@app.route('/api/generate-individual', methods=['POST'])
def generate_individual_timesheet():
    data = request.json or {}
    month_str = data.get('month', datetime.now().strftime('%Y-%m'))
    domain_key = data.get('domain', 'identity')
    member_data = data.get('member')
    member_name_str = data.get('member_name', '')
    public_holidays = data.get('public_holidays', [])
    include_sg_holidays = data.get('auto_sg_holidays', True)
    comment_notes = data.get('comment_notes', "")

    try:
        dt = datetime.strptime(month_str, '%Y-%m')
        year, month = dt.year, dt.month
    except ValueError:
        return jsonify({'error': 'Invalid month format, expected YYYY-MM'}), 400

    template_path = find_template_path(domain_key, TEMPLATE_DIR)
    if not template_path:
        return jsonify({'error': f'Template for domain "{domain_key}" not found.'}), 404

    target_member = None
    if member_data and isinstance(member_data, dict) and member_data.get('name'):
        target_member = member_data
    else:
        try:
            wb = openpyxl.load_workbook(template_path, data_only=True)
            all_m = extract_resources_from_summary(wb)
            wb.close()
            for m in all_m:
                if m.get('name', '').lower() == member_name_str.lower() or m.get('account', '').lower() == member_name_str.lower():
                    target_member = m
                    break
        except Exception:
            pass

    if not target_member:
        # Fallback to Sample Resource if blank or unselected
        target_member = {
            'name': 'Nguyen Van A',
            'account': 'AnNV',
            'team': 'FPT',
            'lead': 'Lead',
            'location': 'Offshore',
            'total_leave': 14,
            'balance_upto': 10
        }

    safe_name = re.sub(r'[^a-zA-Z0-9]', '', target_member.get('account') or target_member.get('name'))
    month_suffix = datetime(year, month, 1).strftime('%b%Y')
    available_domains = scan_available_domain_templates(TEMPLATE_DIR)
    domain_info = available_domains.get(domain_key, {})
    domain_name = domain_info.get('name', domain_key.upper()).replace(' ', '_')

    out_filename = f"FPT_{domain_name}_Timesheet_{month_suffix}_{safe_name}.xlsx"
    out_filepath = os.path.join(OUTPUT_DIR, out_filename)

    generate_domain_timesheet(
        template_path=template_path,
        year=year,
        month=month,
        output_path=out_filepath,
        public_holidays=public_holidays,
        include_sg_holidays=include_sg_holidays,
        custom_members=[target_member],
        comment_notes=comment_notes,
        domain_key=domain_key
    )

    return send_file(
        out_filepath,
        as_attachment=True,
        download_name=out_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("Starting Monthly Timesheet Generator Web Server on http://127.0.0.1:5050 ...")
    app.run(host='127.0.0.1', port=5050, debug=True)
