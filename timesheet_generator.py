import os
import glob
import calendar
import re
import zipfile
from copy import copy
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    import holidays
    HAS_HOLIDAYS_PKG = True
except ImportError:
    HAS_HOLIDAYS_PKG = False

MONTH_NAMES_SHORT = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

DOMAIN_FILE_MAP = {
    'cbg': 'FPT_CBG_CRM-OM_Domain_Timesheet_Template.xlsx',
}

DOMAIN_NAMES = {
    'cbg': 'CBG CRM-OM Domain',
}

# Standard thin border for table content cells
THIN_SIDE = Side(style='thin', color='000000')
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

ALIGN_CENTER_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_CENTER_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT_CENTER = Alignment(horizontal='left', vertical='center')

def get_singapore_holidays(year: int, month: int = None):
    """
    Returns dictionary of Singapore Public Holidays for given year and optional month.
    Key: string YYYY-MM-DD, Value: holiday name.
    """
    results = {}
    if HAS_HOLIDAYS_PKG:
        sg_holidays = holidays.Singapore(years=year)
        for dt, name in sg_holidays.items():
            if month is None or dt.month == month:
                results[dt.strftime("%Y-%m-%d")] = name
    else:
        fallback_2026 = {
            "2026-01-01": "New Year's Day",
            "2026-02-17": "Chinese New Year",
            "2026-02-18": "Chinese New Year",
            "2026-03-21": "Hari Raya Puasa",
            "2026-04-03": "Good Friday",
            "2026-05-01": "Labour Day",
            "2026-05-27": "Hari Raya Haji",
            "2026-05-31": "Vesak Day",
            "2026-06-01": "Vesak Day (observed)",
            "2026-08-09": "National Day",
            "2026-08-10": "National Day (observed)",
            "2026-11-08": "Deepavali",
            "2026-11-09": "Deepavali (observed)",
            "2026-12-25": "Christmas Day"
        }
        for d_str, name in fallback_2026.items():
            dt = datetime.strptime(d_str, "%Y-%m-%d").date()
            if dt.year == year and (month is None or dt.month == month):
                results[d_str] = name
    return results

def get_working_days(year: int, month: int):
    """
    Returns list of date objects for all weekdays (Mon-Fri) in the given year and month.
    Saturdays (5) and Sundays (6) are excluded.
    """
    _, num_days = calendar.monthrange(year, month)
    working_days = []
    for day in range(1, num_days + 1):
        dt = date(year, month, day)
        if dt.weekday() < 5:  # Mon-Fri
            working_days.append(dt)
    return working_days

def clean_domain_name_string(val: str) -> str:
    """
    Cleans domain template filenames or keys to produce standard, human-readable domain names
    without old month tags, 'FPT_' prefixes, or '_Timesheet' / '_Template' / '_May2026' / '_Jun2026' suffixes.
    Example: 'FPT_EBG_CRM_OM_Jun2026.xlsx' -> 'EBG_CRM_OM'
             'FPT_CBG CRM-OM Domain_Timesheet_May2026.xlsx' -> 'CBG_CRM-OM_Domain'
    """
    if not val:
        return 'Domain'
    base = os.path.basename(str(val)).replace('.xlsx', '')
    base = re.sub(r'^FPT[_\s]*', '', base, flags=re.IGNORECASE)
    base = re.sub(r'[_\s]*Timesheet[_\s]*', '_', base, flags=re.IGNORECASE)
    base = re.sub(r'[_\s]*Template[_\s]*', '_', base, flags=re.IGNORECASE)
    base = re.sub(r'[_\s]*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|January|February|March|April|June|July|August|September|October|November|December)\d{4}', '', base, flags=re.IGNORECASE)
    base = base.strip('_ ')
    base = re.sub(r'\s+', '_', base)
    base = re.sub(r'_+', '_', base)
    return base or 'Domain'

def scan_available_domain_templates(template_dir: str = 'timesheets_extracted'):
    """
    Dynamically scans template_dir for all .xlsx files and extracts domain info.
    Returns dictionary mapping domain_key -> metadata.
    """
    domains = {}
    if not os.path.exists(template_dir):
        return domains

    files = glob.glob(os.path.join(template_dir, "*.xlsx"))
    for fpath in sorted(files):
        if not zipfile.is_zipfile(fpath):
            import time
            time.sleep(0.2)
            if not zipfile.is_zipfile(fpath):
                continue
                
        fname = os.path.basename(fpath)
        key = None
        for k, v in DOMAIN_FILE_MAP.items():
            if v.lower() == fname.lower():
                key = k
                break
                
        clean_name = clean_domain_name_string(fname)
        if not key:
            key = clean_name.lower().replace(' ', '_').replace('-', '_')
            
        base_key = key
        counter = 2
        while key in domains and domains[key]['filename'] != fname:
            key = f"{base_key}_{counter}"
            counter += 1

        display_name = DOMAIN_NAMES.get(base_key) or DOMAIN_NAMES.get(key)
        if not display_name:
            display_name = clean_name.replace('_', ' ')
            if counter > 2:
                display_name = f"{display_name} ({counter-1})"
            
        domains[key] = {
            'key': key,
            'name': display_name,
            'clean_domain': clean_name,
            'filename': fname,
            'filepath': fpath
        }
    return domains

def resolve_formula_cell(wb, val):
    """
    Resolves formula cell references like '=Timesheet!G2' or '=Summary!B2' to actual cell values.
    """
    if val is None:
        return ""
    val_str = str(val).strip()
    if val_str.startswith('='):
        m = re.match(r'^=([A-Za-z0-9_ -]+)!([A-Z]+)(\d+)$', val_str, re.IGNORECASE)
        if m:
            target_sheet_name, col_str, row_str = m.group(1), m.group(2), int(m.group(3))
            target_sheet = None
            for name in wb.sheetnames:
                if name.strip().lower() == target_sheet_name.strip().lower():
                    target_sheet = wb[name]
                    break
            if target_sheet:
                col_num = 0
                for char in col_str.upper():
                    col_num = col_num * 26 + (ord(char) - ord('A') + 1)
                resolved_val = target_sheet.cell(row_str, col_num).value
                if resolved_val is not None and not str(resolved_val).startswith('='):
                    return str(resolved_val).strip()
    return val_str

def parse_val_to_float(val, default=None):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    if not val_str or val_str.startswith('='):
        return default
    try:
        return float(val_str)
    except Exception:
        return default

def resolve_cell_value_numeric(wb, val, default=0.0):
    """
    Resolves a cell value or formula string reference (e.g. '=Summary!L3') to a float.
    """
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    if not val_str:
        return default
    if not val_str.startswith('='):
        try:
            return float(val_str)
        except Exception:
            return default

    # If val is a formula like '=Summary!L3'
    m = re.match(r'^=([A-Za-z0-9_ -]+)!([A-Z]+)(\d+)$', val_str, re.IGNORECASE)
    if m:
        target_sheet_name, col_str, row_str = m.group(1), m.group(2), int(m.group(3))
        col_num = 0
        for char in col_str.upper():
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
        for sname in wb.sheetnames:
            if sname.strip().lower() == target_sheet_name.strip().lower():
                target_cell_v = wb[sname].cell(row_str, col_num).value
                return resolve_cell_value_numeric(wb, target_cell_v, default)

    return default

def extract_leave_balances_from_balance_sheet(wb):
    """
    Scans 'Balance Leave' or 'Balance' sheet in wb and returns a dictionary:
    resource_name_lower -> {'total_leave': float, 'leave_balance_upto': float}
    Reads 'Leave Balance in <Month>' (Column J) directly or evaluates formula =G{r}-MAX(I{r}-H{r},0).
    """
    balances = {}
    bal_sheet = None
    for name in wb.sheetnames:
        if 'balance' in name.lower():
            bal_sheet = wb[name]
            break
            
    if not bal_sheet:
        return balances

    header_row = 1
    col_map = {}
    for r in range(1, min(6, bal_sheet.max_row + 1)):
        row_map = {}
        for c in range(1, bal_sheet.max_column + 1):
            val = bal_sheet.cell(r, c).value
            if val:
                row_map[str(val).strip().lower()] = c
        if any(k in row_map for k in ['name', 'resource', 'member', 'full name', 'họ tên', 'nhân viên', 'họ và tên']):
            header_row = r
            col_map = row_map
            break

    name_col = None
    for k in ['name', 'resource', 'member', 'full name', 'họ tên', 'nhân viên', 'họ và tên']:
        if k in col_map:
            name_col = col_map[k]
            break
    if not name_col:
        name_col = 2

    tot_col = None
    in_col = None
    upto_col = None
    annual_col = None
    lieu_col = None

    for h, c in col_map.items():
        if 'total leave' in h: tot_col = c
        elif 'leave balance in' in h: in_col = c
        elif 'leave balance upto' in h: upto_col = c
        elif 'annual leave' in h: annual_col = c
        elif 'days off in lieu' in h: lieu_col = c

    if not tot_col: tot_col = 6 if 'lead' in col_map else 5
    if not upto_col: upto_col = 7 if 'lead' in col_map else 6

    for r in range(header_row + 1, bal_sheet.max_row + 1):
        raw_name = bal_sheet.cell(r, name_col).value
        if not raw_name:
            continue
        res_name = resolve_formula_cell(wb, raw_name)
        if not res_name or str(res_name).startswith('='):
            continue
            
        r_name_clean = str(res_name).strip().lower()
        if r_name_clean in ['total', 'tổng', 'stt', 'no', 'none', 'name', 'resource', '0']:
            continue
            
        tot_val = bal_sheet.cell(r, tot_col).value if tot_col else 14
        tot_num = resolve_cell_value_numeric(wb, tot_val, 14.0)

        # 1. Read direct numeric value in Column J (Leave Balance in Month)
        in_val = bal_sheet.cell(r, in_col).value if in_col else None
        in_float = parse_val_to_float(in_val, None)

        if in_float is not None:
            net_bal = in_float
        else:
            # 2. Evaluate Excel formula =G{r}-MAX(I{r}-H{r},0) by resolving cell references from Balance Leave & Summary sheets
            upto_val = bal_sheet.cell(r, upto_col).value if upto_col else None
            upto_num = resolve_cell_value_numeric(wb, upto_val, 10.0)

            annual_val = bal_sheet.cell(r, annual_col).value if annual_col else None
            annual_num = resolve_cell_value_numeric(wb, annual_val, 0.0)

            lieu_val = bal_sheet.cell(r, lieu_col).value if lieu_col else None
            lieu_num = resolve_cell_value_numeric(wb, lieu_val, 0.0)

            net_bal = max(upto_num - max(annual_num - lieu_num, 0.0), 0.0)

        balances[r_name_clean] = {
            'total_leave': tot_num,
            'leave_balance_upto': net_bal
        }

    return balances

def extract_resources_from_summary(wb):
    """
    Extracts resource list with metadata (Name, Team, Lead, Location, Total Leave, Leave Balance Upto) from Summary & Balance Leave sheets.
    Supports flexible header row scanning, formula resolution, and Vietnamese/English col headers.
    """
    summary_sheet = None
    for name in wb.sheetnames:
        if any(k in name.lower() for k in ['summary', 'danh sách', 'resource', 'nhân viên', 'member']):
            summary_sheet = wb[name]
            break
    if not summary_sheet:
        return []

    sheet = summary_sheet
    resources = []
    bal_data = extract_leave_balances_from_balance_sheet(wb)
    
    header_row = 1
    col_map = {}
    
    # Scan top 5 rows to find header row containing any member name column keyword
    for r in range(1, min(6, sheet.max_row + 1)):
        row_map = {}
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val:
                row_map[str(val).strip().lower()] = c
        if any(k in row_map for k in ['name', 'resource', 'member', 'full name', 'họ tên', 'nhân viên', 'họ và tên', 'staff', 'employee']):
            header_row = r
            col_map = row_map
            break

    name_col = None
    for k in ['name', 'resource', 'member', 'full name', 'họ tên', 'nhân viên', 'họ và tên', 'staff', 'employee']:
        if k in col_map:
            name_col = col_map[k]
            break
    if not name_col:
        name_col = 2  # default Col B

    team_col = None
    for k in ['team', 'phòng ban', 'bộ phận', 'group']:
        if k in col_map:
            team_col = col_map[k]
            break
    if not team_col:
        team_col = 3

    lead_col = None
    for k in ['lead', 'leader', 'quản lý']:
        if k in col_map:
            lead_col = col_map[k]
            break
    if not lead_col:
        lead_col = 4

    loc_col = None
    for k in ['location', 'vị trí', 'onsite/offshore', 'role']:
        if k in col_map:
            loc_col = col_map[k]
            break
    if not loc_col:
        loc_col = 5

    for r in range(header_row + 1, sheet.max_row + 1):
        raw_name = sheet.cell(r, name_col).value
        if not raw_name:
            continue
            
        resolved_name = resolve_formula_cell(wb, raw_name)
        if not resolved_name or resolved_name.startswith('='):
            continue
            
        name_clean = str(resolved_name).strip()
        if name_clean.lower() in ['total', 'tổng', 'stt', 'no', 'none', 'name', 'resource', 'họ và tên', '0']:
            continue
            
        team_val = sheet.cell(r, team_col).value if team_col else ""
        lead_val = sheet.cell(r, lead_col).value if lead_col else ""
        loc_val = sheet.cell(r, loc_col).value if loc_col else "Offshore"
        
        r_name_lower = name_clean.lower()
        member_bal = bal_data.get(r_name_lower, {})
        tot_leave_val = member_bal.get('total_leave', 14)
        leave_upto_val = member_bal.get('leave_balance_upto', 10)

        resources.append({
            'row': r,
            'name': name_clean,
            'team': str(resolve_formula_cell(wb, team_val) or "").strip(),
            'lead': str(resolve_formula_cell(wb, lead_val) or "").strip(),
            'location': str(resolve_formula_cell(wb, loc_val) or "Offshore").strip(),
            'vendor': 'FPT',
            'total_leave': tot_leave_val,
            'leave_balance_upto': leave_upto_val
        })
    return resources

def extract_resources_from_timesheet(sheet_ts):
    """
    Extracts distinct resources, team, vendor from Timesheet sheet if Summary is unavailable or empty.
    Dynamically locates Resource/Name columns across top rows.
    """
    if not sheet_ts:
        return []
        
    wb = sheet_ts.parent
    header_row = 1
    name_col = 7  # Default Col G
    team_col = 8  # Default Col H
    vendor_col = 9  # Default Col I
    
    # Scan top 5 rows for header matching
    for r in range(1, min(6, sheet_ts.max_row + 1)):
        for c in range(1, sheet_ts.max_column + 1):
            val = sheet_ts.cell(r, c).value
            if val:
                val_str = str(val).strip().lower()
                if val_str in ['resource', 'name', 'full name', 'member', 'nhân viên', 'họ và tên', 'staff']:
                    name_col = c
                    header_row = r
                elif val_str in ['team', 'bộ phận', 'phòng ban']:
                    team_col = c
                elif val_str in ['vendor', 'vendor name', 'công ty']:
                    vendor_col = c

    resources_map = {}
    for r in range(header_row + 1, sheet_ts.max_row + 1):
        raw_res = sheet_ts.cell(r, name_col).value
        if not raw_res:
            continue
        res_name = resolve_formula_cell(wb, raw_res)
        if not res_name or res_name.startswith('='):
            continue
            
        res_clean = str(res_name).strip()
        if res_clean.lower() in ['total', 'tổng', 'stt', 'no', 'resource', 'name', 'họ và tên', '0']:
            continue
            
        if res_clean not in resources_map:
            team_val = sheet_ts.cell(r, team_col).value if team_col else ""
            vendor_val = sheet_ts.cell(r, vendor_col).value if vendor_col else "FPT"
            
            resources_map[res_clean] = {
                'name': res_clean,
                'team': str(resolve_formula_cell(wb, team_val) or "").strip(),
                'lead': '',
                'location': 'Offshore',
                'vendor': str(resolve_formula_cell(wb, vendor_val) or "FPT").strip(),
                'total_leave': 14,
                'leave_balance_upto': 10
            }
    return list(resources_map.values())

def extract_resources_from_any_sheet(wb):
    """
    Fallback extractor scanning all sheets for person names if standard sheets return no members.
    """
    for ws in wb.worksheets:
        m_list = extract_resources_from_timesheet(ws)
        if m_list:
            return m_list
    return []


def find_and_read_prev_month_leave_balances(domain_key: str, year: int, month: int, search_dirs=None):
    """
    Looks for the previous month's generated or template timesheet file for domain_key.
    Extracts each resource's calculated ending Leave Balance In month.
    """
    if search_dirs is None:
        base_dir = os.path.dirname(__file__)
        search_dirs = [
            os.path.join(base_dir, 'generated_timesheets'),
            os.path.join(base_dir, 'timesheets_extracted')
        ]
        
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    prev_month_suffix = datetime(prev_year, prev_month, 1).strftime('%b%Y').lower()
    
    prev_balances = {}
    candidate_files = []
    
    clean_dkey = domain_key.lower().replace('_', '').replace(' ', '').replace('-', '')
    for sdir in search_dirs:
        if os.path.exists(sdir):
            for fname in os.listdir(sdir):
                if fname.endswith('.xlsx') and not fname.startswith('~$'):
                    fn_clean = fname.lower().replace('_', '').replace(' ', '').replace('-', '')
                    if prev_month_suffix in fn_clean and clean_dkey in fn_clean:
                        candidate_files.append(os.path.join(sdir, fname))
                        
    if not candidate_files:
        return prev_balances

    target_fpath = candidate_files[0]
    try:
        wb = openpyxl.load_workbook(target_fpath, data_only=True, read_only=True)
        sheet_name = 'Balance Leave' if 'Balance Leave' in wb.sheetnames else ('Leave Balance' if 'Leave Balance' in wb.sheetnames else None)
        sheet_sum_name = 'Summary' if 'Summary' in wb.sheetnames else None
        
        if sheet_name:
            sheet_bal = wb[sheet_name]
            bal_rows = list(sheet_bal.iter_rows(values_only=True))
            col_map = {}
            if len(bal_rows) >= 1:
                for c, v in enumerate(bal_rows[0]):
                    if v: col_map[str(v).strip().lower()] = c + 1
            if len(bal_rows) >= 2:
                for c, v in enumerate(bal_rows[1]):
                    if v: col_map[str(v).strip().lower()] = c + 1
                    
            name_col = col_map.get('name', 2)
            upto_col = None
            in_col = None
            annual_col = None
            lieu_col = None

            for h, c in col_map.items():
                if 'leave balance upto' in h: upto_col = c
                elif 'leave balance in' in h: in_col = c
                elif 'annual leave' in h: annual_col = c
                elif 'days off in lieu' in h: lieu_col = c

            if not upto_col: upto_col = 7 if 'lead' in col_map else 6
            
            sum_leaves_map = {}
            if sheet_sum_name:
                sheet_sum = wb[sheet_sum_name]
                sum_rows = list(sheet_sum.iter_rows(values_only=True))
                sum_col_map = {}
                if len(sum_rows) >= 1:
                    for c, v in enumerate(sum_rows[0]):
                        if v: sum_col_map[str(v).strip().lower()] = c + 1
                sum_name_col = sum_col_map.get('name', 2)
                sum_leaves_col = None
                for h, c in sum_col_map.items():
                    if 'leaves' in h or 'leave' in h:
                        sum_leaves_col = c
                        break
                if sum_leaves_col:
                    for sr in range(1, len(sum_rows)):
                        s_name = sum_rows[sr][sum_name_col - 1]
                        l_val = sum_rows[sr][sum_leaves_col - 1]
                        if s_name:
                            sum_leaves_map[str(s_name).strip().lower()] = float(l_val) if isinstance(l_val, (int, float)) else 0.0

            for r in range(2, len(bal_rows)):
                row_vals = bal_rows[r]
                name_val = row_vals[name_col - 1] if len(row_vals) >= name_col else None
                if name_val and not str(name_val).startswith('='):
                    res_name = str(name_val).strip().lower()
                    
                    # 1. Try reading calculated ending Leave Balance In column first
                    in_val = row_vals[in_col - 1] if (in_col and len(row_vals) >= in_col) else None
                    if isinstance(in_val, (int, float)):
                        prev_balances[res_name] = float(in_val)
                    else:
                        upto_val = row_vals[upto_col - 1] if (upto_col and len(row_vals) >= upto_col) else 10.0
                        upto_num = float(upto_val) if isinstance(upto_val, (int, float)) else 10.0
                        
                        annual_val = row_vals[annual_col - 1] if (annual_col and len(row_vals) >= annual_col) else None
                        if isinstance(annual_val, (int, float)):
                            annual_num = float(annual_val)
                        else:
                            annual_num = sum_leaves_map.get(res_name, 0.0)

                        lieu_val = row_vals[lieu_col - 1] if (lieu_col and len(row_vals) >= lieu_col) else 0.0
                        lieu_num = float(lieu_val) if isinstance(lieu_val, (int, float)) else 0.0

                        prev_balances[res_name] = max(upto_num - annual_num + lieu_num, 0.0)
        wb.close()
    except Exception as e:
        print(f"Error reading prev month balances: {e}")
        
    return prev_balances

def get_row_font_color(work_item_type: str):
    """
    Deprecated: Using Conditional Formatting instead.
    """
    return "000000"

def parse_comment_notes(notes_input, year: int, month: int):
    """
    Parses raw comment table text or list of dicts into structured overrides,
    including leave balances and date-specific OT/Leave entries.
    """
    overrides = {}
    if not notes_input:
        return overrides
        
    lines = []
    if isinstance(notes_input, list):
        for item in notes_input:
            name = item.get('name') or item.get('member_name')
            if name:
                overrides[name.strip().lower()] = item
        return overrides
    elif isinstance(notes_input, str):
        lines = [l.strip() for l in notes_input.strip().split('\n') if l.strip()]

    for line in lines:
        parts = [p.strip() for p in line.split('\t')]
        if len(parts) < 2:
            parts = [p.strip() for p in line.split('|')]
        if len(parts) < 2:
            continue
            
        if 'member' in parts[0].lower() or 'fsoft' in parts[1].lower():
            continue
            
        member_name = parts[0]
        fsoft_acc = parts[1] if len(parts) > 1 else ""
        project = parts[2] if len(parts) > 2 else ""
        note = parts[3] if len(parts) > 3 else parts[-1]
        
        m_name_key = member_name.strip().lower()
        if m_name_key not in overrides:
            overrides[m_name_key] = {
                'name': member_name,
                'fsoft_account': fsoft_acc,
                'project': project,
                'note': note,
                'ot_entries': []
            }
        
        info = overrides[m_name_key]
        
        # Parse leave balances
        m_upto = re.search(r'leave\s*balance\s*upto.*?[về|to|=|:]\s*([\d\.]+)', note, re.IGNORECASE)
        if m_upto:
            info['leave_balance_upto'] = float(m_upto.group(1))
            
        m_in = re.search(r'leave\s*balance\s*in.*?[về|to|=|:]\s*([\d\.]+)', note, re.IGNORECASE)
        if m_in:
            info['leave_balance_in_month'] = float(m_in.group(1))

        # Parse OT & Leave dates ONLY if note contains relevant keywords
        is_ot_or_leave_note = any(k in note.lower() for k in ['ot', 'support', 'weekday', 'weekend', 'ph', 'nghỉ', 'phép', 'update work item', 'line'])
        if is_ot_or_leave_note:
            date_matches = re.finditer(r'(\d{1,2})[-\/\s]*([A-Za-z]{3}|\d{1,2})?', note)
            for m in date_matches:
                d_str = m.group(1)
                m_str = m.group(2)
                
                try:
                    day_num = int(d_str)
                    target_m = month
                    if m_str:
                        if m_str.isdigit():
                            target_m = int(m_str)
                        else:
                            m_lower = m_str.lower()
                            if m_lower in [m_n.lower() for m_n in MONTH_NAMES_SHORT if m_n]:
                                for idx, m_n in enumerate(MONTH_NAMES_SHORT):
                                    if m_n.lower() == m_lower:
                                        target_m = idx
                                        break
                                
                    dt_str = f"{year}-{target_m:02d}-{day_num:02d}"
                    
                    m_hrs = re.search(r'(\d+(?:\.\d+)?)\s*h', note, re.IGNORECASE)
                    hours = float(m_hrs.group(1)) if m_hrs else None
                    
                    note_segment = note.lower()
                    if 'weekend' in note_segment:
                        wtype = 'Weekend support'
                        if hours is None: hours = 8.0
                    elif 'ph support' in note_segment or 'ph ot' in note_segment or ('ph' in note_segment and 'support' in note_segment):
                        wtype = 'PH Support'
                        if hours is None: hours = 8.0
                    elif 'leave' in note_segment or 'nghỉ' in note_segment or 'phép' in note_segment:
                        wtype = 'Leave'
                        if hours is None: hours = 8.0
                    else:
                        wtype = 'Weekday support'
                        if hours is None: hours = 4.0

                    task_desc = "Support activities"
                    m_task = re.search(r'(?:để|for|về)\s+([^,;\(\)]+)', note, re.IGNORECASE)
                    if m_task:
                        task_desc = m_task.group(1).strip()
                        
                    info['ot_entries'].append({
                        'date': dt_str,
                        'work_item_type': wtype,
                        'hours': hours,
                        'task': task_desc,
                        'project': project
                    })
                    break
                except Exception:
                    pass

    return overrides

def apply_work_item_data_validation(sheet_ts):
    """
    Applies Data Validation dropdown list for Work Item Type on Column C.
    Client Rule: Excludes 'Miscellaneous'. Only accepted worktypes allowed.
    """
    dv_work_items = DataValidation(
        type="list",
        formula1='"Project Task,Weekend support,Weekday support,Leave,Public Holiday,PH Support"',
        allow_blank=True
    )
    sheet_ts.add_data_validation(dv_work_items)
    max_r = max(sheet_ts.max_row + 500, 2000)
    dv_work_items.add(f"C2:C{max_r}")

def apply_work_item_conditional_formatting(sheet_ts, max_r):
    """
    Applies Conditional Formatting to dynamically change row font color based on Work Item Type.
    """
    from openpyxl.formatting.rule import FormulaRule
    red_font = Font(color='FF0000')
    brown_font = Font(color='C00000')
    
    red_rule = FormulaRule(formula=['OR($C2="Weekday support", $C2="Weekend support", $C2="PH Support", $C2="Public Holiday")'], font=red_font, stopIfTrue=False)
    brown_rule = FormulaRule(formula=['$C2="Leave"'], font=brown_font, stopIfTrue=False)
    
    cf_range = f"A2:J{max_r}"
    sheet_ts.conditional_formatting.add(cf_range, red_rule)
    sheet_ts.conditional_formatting.add(cf_range, brown_rule)

def format_sheet_dimensions_and_alignment(ws, sheet_name='Timesheet', resources_count=None):
    """
    Standardizes row heights, column widths, thin borders ONLY for table content on Summary & Balance Leave sheets, and cell alignments.
    Timesheet sheet has NO explicit borders applied per user directive.
    """
    ws.row_dimensions[1].height = 28
    if sheet_name in ['Balance Leave', 'Leave Balance']:
        ws.row_dimensions[2].height = 22

    # Determine max_data_col (last column with header text)
    max_data_col = 0
    max_search_col = min(ws.max_column, 50)  # Cap at 50 to prevent 16383 memory explosion
    for c in range(1, max_search_col + 1):
        v1 = ws.cell(1, c).value
        v2 = ws.cell(2, c).value if ws.max_row >= 2 else None
        if v1 is not None or v2 is not None:
            max_data_col = c
    if max_data_col == 0:
        max_data_col = max_search_col

    # Determine max_data_row (content rows only)
    if resources_count is not None and resources_count > 0:
        if sheet_name == 'Summary':
            max_data_row = resources_count + 1
        elif sheet_name in ['Balance Leave', 'Leave Balance']:
            max_data_row = resources_count + 2
        else:
            max_data_row = ws.max_row
    else:
        max_data_row = ws.max_row

    for r in range(1, max_data_row + 1):
        if r > 1:
            ws.row_dimensions[r].height = 22
            
        for c in range(1, max_data_col + 1):
            cell = ws.cell(r, c)
            
            if sheet_name != 'Timesheet':
                cell.border = THIN_BORDER
                    
            if r == 1 or (r == 2 and sheet_name in ['Balance Leave', 'Leave Balance']):
                cell.alignment = ALIGN_CENTER_CENTER_WRAP
            else:
                if sheet_name == 'Timesheet':
                    if c in [1, 2, 10]:  # Month, Date, Hours
                        cell.alignment = ALIGN_CENTER_CENTER
                    else:
                        cell.alignment = ALIGN_LEFT_CENTER
                else:
                    if c in [1, 6, 7, 8, 9, 10, 11, 12, 13]:
                        cell.alignment = ALIGN_CENTER_CENTER
                    else:
                        cell.alignment = ALIGN_LEFT_CENTER
                        
                if sheet_name in ['Balance Leave', 'Leave Balance'] and 3 <= r:
                    if c in [6, 7]:
                        cell.fill = GRAY_FILL

    if sheet_name == 'Timesheet':
        col_widths = {
            1: 14, 2: 15, 3: 22, 4: 28, 5: 36, 6: 30, 7: 24, 8: 16, 9: 14, 10: 18
        }
    elif sheet_name == 'Summary':
        col_widths = {
            1: 8, 2: 24, 3: 16, 4: 16, 5: 14, 6: 22, 7: 20, 8: 20, 9: 20, 10: 20, 11: 16, 12: 14, 13: 16
        }
    else:
        col_widths = {
            1: 8, 2: 24, 3: 16, 4: 16, 5: 14, 6: 22, 7: 26, 8: 18, 9: 18, 10: 28
        }

    for c in range(1, max_data_col + 1):
        col_letter = get_column_letter(c)
        if c in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[c]
        else:
            ws.column_dimensions[col_letter].width = 16

    # Clean up bloated column dimensions from templates to fix slow save performance
    if ws.max_column > max_data_col:
        from openpyxl.utils import column_index_from_string
        keys_to_delete = []
        for col_letter in ws.column_dimensions.keys():
            if column_index_from_string(col_letter) > max_data_col:
                keys_to_delete.append(col_letter)
        for k in keys_to_delete:
            del ws.column_dimensions[k]

def update_summary_sheet_dynamically(sheet_sum, resources):
    """
    Dynamically maps Row 1 header titles to exact formulas and column positions in Summary sheet.
    """
    col_map = {}
    for c in range(1, sheet_sum.max_column + 1):
        v = sheet_sum.cell(1, c).value
        if v:
            col_map[str(v).strip().lower()] = c
            
    name_col_idx = col_map.get('name', 2)
    name_col_let = get_column_letter(name_col_idx)
    
    max_r = sheet_sum.max_row
    if max_r > 1:
        sheet_sum.delete_rows(2, amount=max_r - 1)
        
    for i, res in enumerate(resources):
        r = i + 2
        for h_key, col_idx in col_map.items():
            if h_key == 'no':
                sheet_sum.cell(r, col_idx, '=ROW()-1')
            elif h_key == 'name':
                sheet_sum.cell(r, col_idx, res['name'])
            elif h_key == 'team':
                sheet_sum.cell(r, col_idx, res.get('team', ''))
            elif h_key == 'lead':
                sheet_sum.cell(r, col_idx, res.get('lead', ''))
            elif h_key == 'location':
                sheet_sum.cell(r, col_idx, res.get('location', 'Offshore'))
            elif 'monthly working days' in h_key or (h_key == 'working days (days)' and 'monthly working days' not in col_map):
                sheet_sum.cell(r, col_idx, f'=SUMIFS(Timesheet!J:J,Timesheet!G:G,Summary!{name_col_let}{r})/8')
            elif h_key == 'working days (days)' and 'monthly working days' in col_map:
                sheet_sum.cell(r, col_idx, f'=SUMIFS(Timesheet!J:J,Timesheet!G:G,Summary!{name_col_let}{r})/8')
            elif 'weekend' in h_key:
                sheet_sum.cell(r, col_idx, f'=SUMIFS(Timesheet!J:J,Timesheet!G:G,Summary!{name_col_let}{r},Timesheet!C:C,\"Weekend support\")/8')
            elif 'weekday' in h_key:
                sheet_sum.cell(r, col_idx, f'=SUMIFS(Timesheet!J:J,Timesheet!G:G,Summary!{name_col_let}{r},Timesheet!C:C,\"Weekday support\")/8')
            elif 'ph support' in h_key or 'ph ot' in h_key:
                sheet_sum.cell(r, col_idx, f'=SUMIFS(Timesheet!J:J,Timesheet!G:G,Summary!{name_col_let}{r},Timesheet!C:C,\"PH Support\")/8')
            elif 'total ot' in h_key:
                ot_cols = []
                for h2, c2 in col_map.items():
                    if any(k in h2 for k in ['weekend', 'weekday', 'ph support', 'ph ot']):
                        ot_cols.append(c2)
                if ot_cols:
                    start_let = get_column_letter(min(ot_cols))
                    end_let = get_column_letter(max(ot_cols))
                    sheet_sum.cell(r, col_idx, f'=SUM({start_let}{r}:{end_let}{r})')
            elif 'leaves' in h_key or 'leave' in h_key:
                sheet_sum.cell(r, col_idx, f'=COUNTIFS(Timesheet!G:G,Summary!{name_col_let}{r},Timesheet!C:C,\"Leave\")')
            elif 'public holiday' in h_key:
                sheet_sum.cell(r, col_idx, f'=COUNTIFS(Timesheet!G:G,Summary!{name_col_let}{r},Timesheet!C:C,\"Public Holiday\")')

def update_balance_leave_sheet_dynamically(sheet_bal, resources, parsed_overrides, year, month, sheet_sum=None, domain_key=None):
    """
    Dynamically maps Row 1 & Row 2 header titles to exact formulas and column positions in Balance Leave sheet.
    Inherits Leave Balance Upto from previous month's Leave Balance In column.
    """
    col_map = {}
    for c in range(1, sheet_bal.max_column + 1):
        v1 = sheet_bal.cell(1, c).value
        v2 = sheet_bal.cell(2, c).value
        if v1: col_map[str(v1).strip().lower()] = c
        if v2: col_map[str(v2).strip().lower()] = c

    name_col = col_map.get('name', 2)
    upto_col = None
    in_col = None
    annual_col = None
    lieu_col = None

    for h, c in col_map.items():
        if 'leave balance upto' in h: upto_col = c
        elif 'leave balance in' in h: in_col = c
        elif 'annual leave' in h: annual_col = c
        elif 'days off in lieu' in h: lieu_col = c

    if not upto_col: upto_col = 7 if 'lead' in col_map else 6

    # Extract template default net balances (calculating net ending balance for next month's upto)
    template_balances = {}
    for r in range(3, sheet_bal.max_row + 1):
        name_val = sheet_bal.cell(r, name_col).value
        if name_val and not str(name_val).startswith('='):
            res_name = str(name_val).strip().lower()
            
            in_val = sheet_bal.cell(r, in_col).value if in_col else None
            if isinstance(in_val, (int, float)):
                template_balances[res_name] = float(in_val)
            else:
                upto_val = sheet_bal.cell(r, upto_col).value if upto_col else 10.0
                upto_num = float(upto_val) if isinstance(upto_val, (int, float)) else 10.0
                
                annual_val = sheet_bal.cell(r, annual_col).value if annual_col else 0.0
                annual_num = float(annual_val) if isinstance(annual_val, (int, float)) else 0.0
                
                lieu_val = sheet_bal.cell(r, lieu_col).value if lieu_col else 0.0
                lieu_num = float(lieu_val) if isinstance(lieu_val, (int, float)) else 0.0
                
                template_balances[res_name] = max(upto_num - annual_num + lieu_num, 0.0)

    # Read previous month's leave balances for domain (if exists)
    prev_month_balances = {}
    if domain_key:
        prev_month_balances = find_and_read_prev_month_leave_balances(domain_key, year, month)

    sum_ot_let = 'K'
    sum_leaves_let = 'L'
    if sheet_sum:
        sum_col_map = {}
        for c in range(1, sheet_sum.max_column + 1):
            v = sheet_sum.cell(1, c).value
            if v: sum_col_map[str(v).strip().lower()] = c
        for h2, c2 in sum_col_map.items():
            if 'total ot' in h2: sum_ot_let = get_column_letter(c2)
            elif 'leaves' in h2 or 'leave' in h2: sum_leaves_let = get_column_letter(c2)

    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    prev_month_str = f"{MONTH_NAMES_SHORT[prev_month]} {str(prev_year)[2:]}"
    target_month_str = f"{MONTH_NAMES_SHORT[month]}-{str(year)[2:]}"
    
    for c in range(1, sheet_bal.max_column + 1):
        val = sheet_bal.cell(1, c).value
        if val:
            val_str = str(val)
            if 'Leave Balance upto' in val_str:
                sheet_bal.cell(1, c, f"Leave Balance upto {prev_month_str}")
            elif 'Leave Balance in' in val_str:
                sheet_bal.cell(1, c, f"Leave Balance in {target_month_str} (days)")
            elif isinstance(val, (datetime, date)):
                sheet_bal.cell(1, c, datetime(year, month, 1))

    max_r_bal = sheet_bal.max_row
    if max_r_bal > 2:
        sheet_bal.delete_rows(3, amount=max_r_bal - 2)
        
    upto_col_let = 'G'
    lieu_col_let = 'H'
    annual_col_let = 'I'

    for h_key, col_idx in col_map.items():
        if 'leave balance upto' in h_key: upto_col_let = get_column_letter(col_idx)
        elif 'days off in lieu' in h_key: lieu_col_let = get_column_letter(col_idx)
        elif 'annual leave' in h_key: annual_col_let = get_column_letter(col_idx)

    for i, res in enumerate(resources):
        r = i + 3
        res_name = res['name']
        member_override = parsed_overrides.get(res_name.lower(), {})
        
        # Determine Leave Balance Upto: Priority 1: Comment Note override -> Priority 2: Previous Month File -> Priority 3: Member custom res -> Priority 4: Template default -> Priority 5: 10
        if 'leave_balance_upto' in member_override:
            leave_upto = member_override['leave_balance_upto']
        elif res_name.lower() in prev_month_balances:
            leave_upto = prev_month_balances[res_name.lower()]
        elif 'leave_balance_upto' in res and res['leave_balance_upto'] is not None:
            leave_upto = res['leave_balance_upto']
        elif res_name.lower() in template_balances:
            leave_upto = template_balances[res_name.lower()]
        else:
            leave_upto = res.get('leave_balance_upto', 10)

        total_leave = res.get('total_leave', 14)
        
        for h_key, col_idx in col_map.items():
            if h_key == 'no':
                sheet_bal.cell(r, col_idx, '=ROW()-2')
            elif h_key == 'name':
                sheet_bal.cell(r, col_idx, res_name)
            elif h_key == 'team':
                sheet_bal.cell(r, col_idx, res.get('team', ''))
            elif h_key == 'lead':
                sheet_bal.cell(r, col_idx, res.get('lead', ''))
            elif h_key == 'location':
                sheet_bal.cell(r, col_idx, res.get('location', 'Offshore'))
            elif 'total leave in' in h_key:
                sheet_bal.cell(r, col_idx, total_leave)
            elif 'leave balance upto' in h_key:
                sheet_bal.cell(r, col_idx, leave_upto)
            elif 'days off in lieu' in h_key:
                sheet_bal.cell(r, col_idx, f'=Summary!{sum_ot_let}{r-1}')
            elif 'annual leave' in h_key:
                sheet_bal.cell(r, col_idx, f'=Summary!{sum_leaves_let}{r-1}')
            elif 'leave balance in' in h_key:
                sheet_bal.cell(r, col_idx, f'={upto_col_let}{r}-MAX({annual_col_let}{r}-{lieu_col_let}{r},0)')

def generate_domain_timesheet(template_path: str, year: int, month: int, output_path: str, public_holidays: list = None, include_sg_holidays: bool = True, custom_members: list = None, comment_notes = None, domain_key: str = None):
    """
    Generates a new timesheet Excel file for the target year and month based on template_path.
    Preserves all formulas, formatting, applies thin table borders ONLY to actual content on Summary & Balance Leave, and excludes Sat/Sun.
    """
    if public_holidays is None:
        public_holidays = []
    
    ph_dates = set()
    ph_dict = {}
    
    if include_sg_holidays:
        sg_hols = get_singapore_holidays(year, month)
        for d_str, name in sg_hols.items():
            dt_obj = datetime.strptime(d_str, "%Y-%m-%d").date()
            ph_dates.add(dt_obj)
            ph_dict[dt_obj] = name

    for ph in public_holidays:
        if isinstance(ph, str):
            try:
                dt_obj = datetime.strptime(ph.strip(), "%Y-%m-%d").date()
                ph_dates.add(dt_obj)
                if dt_obj not in ph_dict:
                    ph_dict[dt_obj] = "Public Holiday"
            except ValueError:
                pass
        elif isinstance(ph, (date, datetime)):
            dt_obj = ph if isinstance(ph, date) else ph.date()
            ph_dates.add(dt_obj)
            if dt_obj not in ph_dict:
                ph_dict[dt_obj] = "Public Holiday"
            
    parsed_overrides = parse_comment_notes(comment_notes, year, month)
    
    # Load template preserving formulas
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    # Extract / set resources list
    if custom_members and len(custom_members) > 0:
        resources = custom_members
    else:
        summary_resources = extract_resources_from_summary(wb)
        if summary_resources:
            resources = summary_resources
        else:
            resources = extract_resources_from_timesheet(wb['Timesheet']) if 'Timesheet' in wb.sheetnames else []

    res_count = len(resources)

    # 1. Process Timesheet Sheet
    if 'Timesheet' in wb.sheetnames:
        sheet_ts = wb['Timesheet']
        working_days = get_working_days(year, month)
        month_first_day = datetime(year, month, 1)
        
        row2_styles = {}
        for c in range(1, 11):
            cell = sheet_ts.cell(2, c)
            row2_styles[c] = {
                'font': copy(cell.font) if cell.font else None,
                'fill': copy(cell.fill) if cell.fill else None,
                'border': copy(cell.border) if cell.border else None,
                'alignment': copy(cell.alignment) if cell.alignment else None,
                'number_format': cell.number_format
            }
            
        max_r = sheet_ts.max_row
        if max_r > 1:
            sheet_ts.delete_rows(2, amount=max_r - 1)
            
        current_row = 2
        font_cache = {}
        def write_ts_row(r_idx, dt_val, wtype, surf_id, proj_val, task_val, res_name, res_team, res_vendor, hours_val):
            sheet_ts.cell(r_idx, 1, month_first_day)
            sheet_ts.cell(r_idx, 2, dt_val)
            sheet_ts.cell(r_idx, 3, wtype)
            sheet_ts.cell(r_idx, 4, surf_id)
            sheet_ts.cell(r_idx, 5, proj_val)
            sheet_ts.cell(r_idx, 6, task_val)
            sheet_ts.cell(r_idx, 7, res_name)
            sheet_ts.cell(r_idx, 8, res_team)
            sheet_ts.cell(r_idx, 9, res_vendor)
            sheet_ts.cell(r_idx, 10, hours_val)

            wtype_lower = str(wtype).lower().strip()
            if any(k in wtype_lower for k in ["public holiday", "ph support", "weekend", "weekday"]):
                row_color = "FF0000"
            elif "leave" in wtype_lower:
                row_color = "C00000"
            else:
                row_color = "000000"

            for c in range(1, 11):
                cell = sheet_ts.cell(r_idx, c)
                st = row2_styles.get(c)
                if st:
                    if st['fill']: cell.fill = st['fill']
                    font_key = (c, row_color)
                    if font_key not in font_cache:
                        base_font = st['font']
                        font_cache[font_key] = Font(
                            name=base_font.name if base_font else "Calibri",
                            size=base_font.size if base_font else 11,
                            bold=base_font.bold if base_font else False,
                            italic=base_font.italic if base_font else False,
                            color=row_color
                        )
                    cell.font = font_cache[font_key]
            sheet_ts.cell(r_idx, 1).number_format = 'mmm-yy'
            sheet_ts.cell(r_idx, 2).number_format = 'd-mmm-yy'

        for res in resources:
            res_name = res['name']
            res_team = res.get('team', '')
            res_vendor = res.get('vendor', 'FPT')
            
            member_override = parsed_overrides.get(res_name.lower(), {})
            ot_entries = member_override.get('ot_entries', [])

            ot_by_date = {}
            for ot_item in ot_entries:
                d_key = ot_item.get('date')
                if d_key:
                    if d_key not in ot_by_date:
                        ot_by_date[d_key] = []
                    ot_by_date[d_key].append(ot_item)
            
            # 1. Mon-Fri Working Days
            for wday in working_days:
                wday_dt = datetime(wday.year, wday.month, wday.day)
                wday_str = wday.strftime('%Y-%m-%d')
                is_ph = wday in ph_dates
                day_ot_list = ot_by_date.get(wday_str, [])

                leave_entry = next((e for e in day_ot_list if e['work_item_type'] == 'Leave'), None)

                # Base / Standard Row
                if leave_entry:
                    wtype = "Leave"
                    surf_id = "Leave"
                    proj_val = leave_entry.get('project', '')
                    task_val = leave_entry.get('task', 'Annual Leave')
                    hours_val = 0  # Client Rule: Leave Actual Time MUST BE 0

                    # If partial day leave (e.g. 4h work + leave), write Project Task line first
                    work_hrs = leave_entry.get('work_hours')
                    if isinstance(work_hrs, (int, float)) and work_hrs > 0:
                        write_ts_row(current_row, wday_dt, "Project Task", "", proj_val, "Production support", res_name, res_team, res_vendor, work_hrs)
                        current_row += 1
                elif is_ph:
                    wtype = "Public Holiday"
                    surf_id = "PH"
                    proj_val = ph_dict.get(wday, "Public Holiday")
                    task_val = ""
                    hours_val = 0
                else:
                    wtype = "Project Task"
                    surf_id = ""
                    proj_val = ""
                    task_val = "Production support"
                    hours_val = 8

                write_ts_row(current_row, wday_dt, wtype, surf_id, proj_val, task_val, res_name, res_team, res_vendor, hours_val)
                current_row += 1

                # Extra OT Rows for this day (e.g. Weekday support, PH Support)
                non_leave_ot = [e for e in day_ot_list if e['work_item_type'] != 'Leave']
                for ot_item in non_leave_ot:
                    ot_wtype = ot_item['work_item_type']
                    ot_hours = ot_item.get('hours', 4)
                    ot_task = ot_item.get('task', 'Support activities')
                    ot_proj = ot_item.get('project', '')

                    write_ts_row(current_row, wday_dt, ot_wtype, "Task", ot_proj, ot_task, res_name, res_team, res_vendor, ot_hours)
                    current_row += 1

            # 2. Weekend OT Rows (Sat/Sun)
            for ot_item in ot_entries:
                ot_date_str = ot_item.get('date')
                if ot_date_str:
                    try:
                        dt_obj = datetime.strptime(ot_date_str, '%Y-%m-%d').date()
                        if dt_obj.weekday() >= 5 and dt_obj.year == year and dt_obj.month == month:
                            wk_dt = datetime(dt_obj.year, dt_obj.month, dt_obj.day)
                            ot_wtype = ot_item['work_item_type']
                            ot_hours = ot_item.get('hours', 8)
                            ot_task = ot_item.get('task', 'Support activities')
                            ot_proj = ot_item.get('project', '')

                            write_ts_row(current_row, wk_dt, ot_wtype, "Task", ot_proj, ot_task, res_name, res_team, res_vendor, ot_hours)
                            current_row += 1
                    except Exception:
                        pass

        apply_work_item_data_validation(sheet_ts)
        apply_work_item_conditional_formatting(sheet_ts, current_row - 1)
        format_sheet_dimensions_and_alignment(sheet_ts, 'Timesheet', res_count)

    # 2. Update Summary Sheet Dynamically
    sheet_sum = wb['Summary'] if 'Summary' in wb.sheetnames else None
    if sheet_sum:
        update_summary_sheet_dynamically(sheet_sum, resources)
        format_sheet_dimensions_and_alignment(sheet_sum, 'Summary', res_count)

    # 3. Update Balance Leave Sheet Dynamically with Previous Month Inherited Balances
    leave_sheet_name = 'Balance Leave' if 'Balance Leave' in wb.sheetnames else ('Leave Balance' if 'Leave Balance' in wb.sheetnames else None)
    if leave_sheet_name:
        sheet_bal = wb[leave_sheet_name]
        update_balance_leave_sheet_dynamically(sheet_bal, resources, parsed_overrides, year, month, sheet_sum, domain_key)
        format_sheet_dimensions_and_alignment(sheet_bal, leave_sheet_name, res_count)

    # 4. Automatically clear active column filters across all worksheets
    clear_all_worksheet_filters(wb)

    # 5. Trim unused trailing rows & clean cell dimensions to optimize file size
    trim_and_optimize_workbook(wb, max_ts_row=current_row - 1, res_count=res_count)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    wb.close()
    return output_path

def trim_and_optimize_workbook(wb, max_ts_row=None, res_count=None):
    """
    Scans all worksheets in the workbook, removes unused blank rows & columns past actual content,
    and cleans up unneeded row_dimensions to dramatically reduce generated Excel file size.
    """
    for ws in wb.worksheets:
        title_lower = ws.title.lower()

        # 1. Determine actual last valid content row for this sheet
        actual_last_row = 1

        if title_lower == 'timesheet' and max_ts_row:
            actual_last_row = max_ts_row
        elif title_lower == 'summary' and res_count is not None:
            actual_last_row = 1 + res_count  # 1 header row + N members
        elif title_lower in ['balance leave', 'leave balance'] and res_count is not None:
            actual_last_row = 2 + res_count  # 2 header rows + N members
        else:
            # General content row scan
            for r in range(ws.max_row, 0, -1):
                has_content = False
                for c in range(1, min(ws.max_column + 1, 35)):
                    val = ws.cell(r, c).value
                    if val is not None and str(val).strip() != '':
                        has_content = True
                        break
                if has_content:
                    actual_last_row = r
                    break

        # 2. Delete empty trailing rows
        if ws.max_row > actual_last_row:
            rows_to_delete = ws.max_row - actual_last_row
            ws.delete_rows(actual_last_row + 1, rows_to_delete)

        # 3. Remove row_dimensions for deleted rows
        invalid_rows = [r for r in list(ws.row_dimensions.keys()) if r > actual_last_row]
        for r in invalid_rows:
            del ws.row_dimensions[r]

def clear_all_worksheet_filters(wb):
    """
    Clears active column filtering criteria and unhides all rows on all worksheets
    so generated Excel workbooks never open with active column filters applied.
    """
    from openpyxl.utils import get_column_letter

    for ws in wb.worksheets:
        # 1. Unhide all rows that might have been hidden by active filters
        for r_idx in range(1, ws.max_row + 1):
            if r_idx in ws.row_dimensions:
                ws.row_dimensions[r_idx].hidden = False

        # 2. Reset column filtering criteria while preserving header dropdown range
        if ws.auto_filter:
            try:
                ws.auto_filter.filterColumn.clear()
            except Exception:
                pass
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

def find_template_path(domain_key: str, template_dir: str = 'timesheets_extracted', allow_fallback: bool = True):
    """
    Finds template file path for a domain key by checking exact keys, substrings, tokens, or filenames.
    Only falls back to default base template if allow_fallback=True.
    """
    if not domain_key:
        return None

    domains = scan_available_domain_templates(template_dir)
    domain_key_clean = str(domain_key).strip().lower()

    # 1. Exact key match in scanned domains
    if domain_key_clean in domains:
        return domains[domain_key_clean]['filepath']

    # 2. Key match in predefined DOMAIN_FILE_MAP
    if domain_key_clean in DOMAIN_FILE_MAP:
        fname = DOMAIN_FILE_MAP[domain_key_clean]
        fpath = os.path.join(template_dir, fname)
        if os.path.exists(fpath):
            return fpath

    # 3. Substring match against scanned keys or filenames
    for k, info in domains.items():
        if domain_key_clean in k or k in domain_key_clean:
            return info['filepath']
        if domain_key_clean in info['filename'].lower():
            return info['filepath']

    # 4. Tokenized search on filename in template_dir
    tokens = [t for t in re.split(r'[_ -]+', domain_key_clean) if t and t not in ['fpt', 'domain', 'timesheet', 'template', 'xlsx']]
    all_files = glob.glob(os.path.join(template_dir, "*.xlsx"))
    
    if tokens and all_files:
        best_match = None
        best_score = 0
        for fpath in all_files:
            fname_lower = os.path.basename(fpath).lower()
            score = sum(1 for tok in tokens if tok in fname_lower)
            if score > best_score:
                best_score = score
                best_match = fpath
        if best_match and best_score > 0:
            return best_match

    # 5. Wildcard glob matching
    pattern = os.path.join(template_dir, f"*{domain_key_clean}*.xlsx")
    matched = glob.glob(pattern)
    if matched:
        return matched[0]

    # 6. Fallback only if allow_fallback is True
    if allow_fallback:
        if domains:
            first_key = list(domains.keys())[0]
            return domains[first_key]['filepath']

        all_xlsx = glob.glob(os.path.join(template_dir, "*.xlsx"))
        if all_xlsx:
            return sorted(all_xlsx)[0]

    return None

